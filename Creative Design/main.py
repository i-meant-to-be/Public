from datetime import datetime
from math import floor, log
import openpyxl

list_genre = {
    1: ("액션", "모험", "애니메이션", "코미디", "SF", "범죄 및 스릴러", "판타지"),
    2: ("드라마", "전쟁", "로맨스", "뮤지컬"),
    3: ("그 외")
}
weight = (5, 2.5, 1)

if __name__ == "__main__":
    # 파일 불러오기
    wb = openpyxl.load_workbook(filename = "data.xlsx", data_only = True)
    ws = wb["Sheet"]

    # 상업성 계산
    for y in range(6, 7):
        try:
            score = 0
            factor = 0
            
            # 장르 (중대함)
            if ws["B" + str(y)].value in list_genre[1]: factor = 50
            elif ws["B" + str(y)].value in list_genre[2]: factor = 25
            else: factor = 5
            score += weight[0] * factor

            # 배우 (중대함)
            if ws["C" + str(y)].value == "성우": factor = 20
            else: factor = (ws["C" + str(y)].value + ws["D" + str(y)].value + ws["E" + str(y)].value + ws["F" + str(y)].value + ws["G" + str(y)].value) / 50000
            score += weight[0] * factor

            # 특정 IP와의 연관성 (중대함)
            if type(ws["H" + str(y)].value) == int:
                T = datetime.now().year - ws["H" + str(y)].value
                prev = ws["I" + str(y)].value
                S_prev = 0
                temp = 0
                for num in range(ord("J"), ord("O")):
                    if type(ws[chr(num) + str(y)].value) == int:
                        S_prev += ws[chr(num) + str(y)].value
                        temp += 1
                S_prev = S_prev / temp / 50000
                S_loyalty = T * min([prev, 5]) * 10 / 9
                factor = S_prev + S_loyalty
                score += weight[0] * factor
        

            # 개봉 시기
            N_curr = ws["O" + str(y)].value
            T = floor((datetime.now() - ws["P" + str(y)].value).total_seconds() // 86400)
            if N_curr < 20: factor = N_curr - N_curr / 30 * T
            else: factor = N_curr - N_curr / 60 * T
            score += weight[1] * factor

            # 감독
            factor = (ws["U" + str(y)].value + ws["Q" + str(y)].value + ws["R" + str(y)].value + ws["S" + str(y)].value + ws["T" + str(y)].value) / 50000
            score += weight[0] * factor

            # 평점 및 언급량
            def __deltaC(A: int, B: int, C: list) -> int:
                return ((C[A] - C[B]) / C[B]) * 100
            C = (ws["V" + str(y)].value,  ws["W" + str(y)].value, ws["X" + str(y)].value, ws["Y" + str(y)].value)
            R = ws["Z" + str(y)].value
            factor = log(R + 1, 2) * (0.2 * __deltaC(3, 2, C) + 0.3 * __deltaC(2, 1, C) + 0.5 * __deltaC(1, 0, C))
            score += weight[1] * factor

            # 상업성 점수 출력
            ws["AA" + str(y)].value = score

            # 결과 출력
            print(f"# Succesfully processed on row {y}.")
        except:
            print(f"# An exception has raised on row {y}. Keep processing for next row.")
    # 엑셀 문서 닫기
    wb.save("data.xlsx")
    wb.close()
    print("# Scripts end.")